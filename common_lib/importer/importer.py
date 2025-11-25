# common_lib/import_/importer.py

import inspect
from typing import Optional, Dict, List

import pandas as pd
from openpyxl import load_workbook

from common_lib.export.constants import DataConstant
from common_lib.importer.models import import_configuration
from common_lib.logging_utils import setup_logger  


class FileImport:
    def __init__(self, file_name: str, template_name: str):
        self._template = import_configuration.import_template[template_name]
        self._file_name = file_name
        self._sep = self._template.separate
        self.logger = setup_logger("IMPORT")
        self._note = ''
        self._endColumnNumer: Optional[int] = None

    def __enter__(self):
        return self

    def __del__(self):
        self._clear_attributes()

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.logger.info(f"[FileImport] Exiting context for file='{self._file_name}'")
        self._clear_attributes()

    def _clear_attributes(self):
        attributes = [
            '_file_name', '_note', '_sep',
            '_template', 'logger', '_endColumnNumer'
        ]
        for attr in attributes:
            setattr(self, attr, None)

    # -------------------- properties --------------------
    @property
    def note(self) -> str:
        return self._note

    @property
    def template(self):
        return self._template

    @property
    def file_name(self):
        return self._file_name

    @property
    def endColumnNumer(self):
        return self._endColumnNumer

    # -------------------- helpers --------------------
    def configEndRow(self, max_row: int) -> int:
        end_row = self.template.end_row or 0
        if end_row <= 0:
            end_row = max_row
        return end_row

    def configEndCol(self, max_col: int) -> int:
        end_col = self.template.end_col or 0
        if end_col <= 0:
            end_col = max_col
        self._endColumnNumer = end_col
        return end_col

    def __readToEmptyCol(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            self.logger.info("[FileImport][__readToEmptyCol] DataFrame is empty.")
            return df

        if self.endColumnNumer is not None and 0 < self.endColumnNumer <= len(df.columns):
            df = df.iloc[:, : self.endColumnNumer + 1]
        elif self.template.header is not None and 0 < len(self.template.header) <= len(df.columns):
            df = df.iloc[:, : len(self.template.header)]
        else:
            empty_col_index = df.columns[df.iloc[0].isnull()].min()
            if pd.notna(empty_col_index):
                df = df.iloc[:, : empty_col_index]

        return df

    def __readToEmptyRow(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            self.logger.info("[FileImport][__readToEmptyRow] DataFrame is empty.")
            return df

        if self.template.is_read_to_empty_row:
            empty_row_index = df[df.isnull().all(axis=1)].index.min()
            if pd.notna(empty_row_index):
                df = df.iloc[:empty_row_index]
        else:
            df = df.dropna(how='all')
        return df

    def __df_replace_header(self, input_header: List[str], config_header: Dict[str, str]) -> List[str]:
        if len(input_header) != len(config_header):
            self.logger.warning(
                f"[FileImport][__df_replace_header] Header length mismatch: "
                f"file={len(input_header)}, config={len(config_header)}"
            )
            unmapped = [h for h in input_header if h not in config_header]
            if unmapped:
                self.logger.info(f"[FileImport][__df_replace_header] Unmapped headers: {unmapped}")
            return input_header

        mapped_header = [config_header.get(col, col) for col in input_header]
        return mapped_header if len(mapped_header) == len(input_header) else input_header

    def __replace_none_with_null(self, value):
        if pd.isna(value):
            return '<NULL>'
        return value

    # -------------------- Excel (openpyxl) --------------------
    def __readExcelToPd(self, file_name: str) -> pd.DataFrame:
        self.logger.info(
            f"[FileImport][__readExcelToPd] Reading Excel file: {file_name}, "
            f"template separate = '{self._sep}', template='{self._template}'"
        )
        wb = load_workbook(file_name)
        sheet = wb.active

        data = []
        for row in sheet.iter_rows(
            min_row=self.template.start_row,
            max_row=self.configEndRow(sheet.max_row),
            min_col=self.template.start_col,
            max_col=self.configEndCol(sheet.max_column),
        ):
            data.append([cell.value for cell in row])

        df = pd.DataFrame(data)
        df = self.__readToEmptyRow(df)
        df = self.__readToEmptyCol(df)
        df = df.map(self.__replace_none_with_null)

        header = self.template.header
        if not header:
            header = [cell.value for cell in sheet[1]]

        mapped_table_header = self.template.column_mapping
        if mapped_table_header:
            header = self.__df_replace_header(header, mapped_table_header)

        if header:
            header = [item.upper() for item in header]
            if len(header) < df.shape[1]:
                df = df.iloc[:, : len(header)]
            elif len(header) > df.shape[1]:
                for i in range(len(header) - df.shape[1]):
                    df[f'_EXTRA_{i}'] = None
            df.columns = header

        self.logger.debug(f"[FileImport][__readExcelToPd] DataFrame shape after read: {df.shape}")
        self.logger.debug(f"[FileImport][__readExcelToPd] Header applied: {header}")
        return df

    # -------------------- CSV/TXT --------------------
    def __safe_read_csv(self, file_name: str, **kwargs) -> pd.DataFrame:
        encodings_to_try = []

        if self._template.file_encoding:
            encodings_to_try.append(self._template.file_encoding)

        encodings_to_try += [
            'ascii',
            'utf-8',
            'utf-8-sig',
            'ISO-8859-1',
            'windows-1252',
            'latin1',
            'utf-16',
            'utf-32',
        ]

        for enc in encodings_to_try:
            try:
                df = pd.read_csv(file_name, encoding=enc, **kwargs)
                self.logger.info(f"[FileImport][__safe_read_csv] Successfully read with encoding: {enc}")
                return df
            except UnicodeDecodeError as uerr:
                self.logger.warning(
                    f"[FileImport][__safe_read_csv] UnicodeDecodeError with encoding '{enc}': {uerr}"
                )
            except Exception as ex:
                self.logger.warning(
                    f"[FileImport][__safe_read_csv] Exception with encoding '{enc}': {ex}"
                )

        raise Exception(
            f"[{self.__class__.__name__}][{inspect.currentframe().f_code.co_name}][{file_name}] encoding not support."
        )

    def __safe_assign_header(self, df: pd.DataFrame, header: List[str], strict: bool = True) -> pd.DataFrame:
        if not header:
            return df

        if strict:
            if len(header) != df.shape[1]:
                self.logger.warning(
                    f"[FileImport][__safe_assign_header] Header length mismatch: "
                    f"config={len(header)}, df={df.shape[1]}"
                )
        else:
            if len(header) < df.shape[1]:
                df = df.iloc[:, : len(header)]
            elif len(header) > df.shape[1]:
                for i in range(len(header) - df.shape[1]):
                    df[f'_EXTRA_{i}'] = None
            self.logger.info("[FileImport][__safe_assign_header] Adjusted DataFrame columns to match header length.")

        df.columns = header
        self.logger.debug(f"[FileImport][__safe_assign_header] Header assigned: {header}")
        return df

    def __readDelimitedFileToPd(self, file_name: str, sep: str = ',') -> pd.DataFrame:
        self.logger.info(
            f"[FileImport][__readDelimitedFileToPd] Reading delimited file: {file_name}, "
            f"template separate = '{self._sep}', input sep='{sep}', template='{self._template}'"
        )

        skiprows = max(self.template.start_row - 1, 0)
        end_row = self.template.end_row or 0
        nrows = None if end_row <= 0 else end_row - self.template.start_row + 1

        start_col = self.template.start_col - 1
        end_col = self.template.end_col - 1 if self.template.end_col and self.template.end_col > 0 else None
        usecols = range(start_col, end_col + 1) if end_col is not None else None

        self.logger.debug(
            f"[FileImport][__readDelimitedFileToPd] skiprows={skiprows}, nrows={nrows}, usecols={usecols}"
        )

        df = self.__safe_read_csv(
            file_name,
            sep=sep,
            header=None,
            skiprows=skiprows,
            nrows=nrows,
            usecols=usecols,
        )

        df = df.map(self.__replace_none_with_null)

        header = self.template.header
        mapped_table_header = self.template.column_mapping
        if mapped_table_header and header:
            header = self.__df_replace_header(header, mapped_table_header)

        if header:
            header = [item.upper() for item in header]
            df = self.__safe_assign_header(df, header)

        self.logger.debug(f"[FileImport][__readDelimitedFileToPd] DataFrame shape: {df.shape}")
        return df

    # -------------------- type helpers --------------------
    @property
    def _is_excel_file(self) -> bool:
        return self.template.file_extension in (
            DataConstant.RPT_EXCEL_FILE_TYPE,
            DataConstant.RPT_EXCEL_FILE_TYPE_OLD,
        )

    @property
    def _is_csv_file(self) -> bool:
        return self.template.file_extension == DataConstant.RPT_CSV_FILE_TYPE

    @property
    def _is_txt_file(self) -> bool:
        return self.template.file_extension == DataConstant.RPT_TEXT_FILE_TYPE

    # -------------------- logging summary --------------------
    def __log_summary(self, df: pd.DataFrame, label: str = ""):
        self.logger.info("[FileImport][__log_summary] START ===================================================")
        if df.empty:
            self.logger.warning(f"[FileImport][__log_summary] {label} DataFrame is empty.")
            return
        self.logger.info(f"[FileImport][__log_summary] {label} Shape: {df.shape}")
        self.logger.info(f"[FileImport][__log_summary] {label} Dtypes: {df.dtypes.to_dict()}")
        self.logger.info("[FileImport][__log_summary] ===================================================== END")

    # -------------------- public API --------------------
    def read_file_to_df(self) -> pd.DataFrame:
        df = pd.DataFrame()
        label = "[START]"
        try:
            if self._is_excel_file:
                df = self.__readExcelToPd(self.file_name)
                label = "[SUCCESS]"
            elif self._is_csv_file:
                if not self._sep or self._sep == 'None':
                    self._sep = ','
                df = self.__readDelimitedFileToPd(self.file_name, sep=self._sep)
                label = "[SUCCESS]"
            elif self._is_txt_file:
                df = self.__readDelimitedFileToPd(self.file_name, sep=self._sep)
                label = "[SUCCESS]"
            else:
                self.logger.warning(
                    f"Unsupported file '{self.file_name}' with format: '{self.template.file_extension}'"
                )
                label = "[UNSUPPORTED]"
        except FileNotFoundError:
            self.logger.error(
                f"[{self.__class__.__name__}][{inspect.currentframe().f_code.co_name}] "
                f"File not found: '{self.file_name}'"
            )
            label = "[ERROR]"
        except PermissionError:
            self.logger.error(
                f"[{self.__class__.__name__}][{inspect.currentframe().f_code.co_name}] "
                f"Permission denied: '{self.file_name}'"
            )
            label = "[ERROR]"
        except Exception as e:
            self.logger.error(
                f"[{self.__class__.__name__}][{inspect.currentframe().f_code.co_name}] "
                f"Read file '{self.file_name}' is Error: {e}"
            )
            label = "[ERROR]"

        if df.empty:
            self.logger.warning("[FileImport][read_file_to_df] File Empty")

        self.__log_summary(df, label=label)
        return df