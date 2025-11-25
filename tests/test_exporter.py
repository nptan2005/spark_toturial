import os
from pathlib import Path
import pandas as pd
import pytest
from openpyxl import load_workbook

from common_lib.export.models import load_export_config
from common_lib.export.exporter import DataExporter

CONFIG_PATH = "common_lib/configurable/export_config.yaml"


@pytest.fixture
def export_config():
    """Load export_config.yaml once."""
    return load_export_config(CONFIG_PATH)


@pytest.fixture
def sample_data():
    """Fake data cho test."""
    df1 = pd.DataFrame({
        "NO": [1, 2],
        "RPT_SESSION": ["20250101", "20250102"],
        "ACQ_BANK": ["VCB", "VTB"],
        "NUM_OF_TXN": [10, 20],
    })

    df2 = pd.DataFrame({
        "NO": [1],
        "RPT_SESSION": ["20250101"],
        "MERCHANT_NAME": ["Shop ABC"],
        "MID": ["M123"],
        "TID": ["T555"],
        "NUM_OF_TXN": [5],
    })

    return [df1, df2]


def test_exporter_excel_creation(tmp_path, export_config, sample_data):
    """
    Test full cycle:
    - Tạo DataExporter
    - Export Excel
    - Kiểm tra file tồn tại
    - Validate sheet & title & header mapping
    """

    export_name = ["ACQ_BANK", "MERCHANT"]
    output_file = tmp_path / "export_test"

    exporter = DataExporter(
        export_data=sample_data,
        export_name=export_name,
        file_name=str(output_file),
        template_key="ACQ_POS_ACCOUNTING_REPORT",
        export_config=export_config,
    )

    exporter.to_file()

    # ======== CHECK 1: file tồn tại ========
    generated_file = tmp_path / "export_test.xlsx"
    assert generated_file.exists(), "Excel file không được tạo!"

    # ======== CHECK 2: Có đúng sheet ========
    wb = load_workbook(generated_file)
    sheet_names = wb.sheetnames

    assert "Tổng hợp ACQ" in sheet_names
    assert "Tổng hợp DVCNT" in sheet_names

    # ======== CHECK 3: Title merged cell ========
    ws = wb["Tổng hợp ACQ"]
    title_cell = ws.cell(row=1, column=1).value
    assert "TỔNG HỢP THANH QUYẾT TOÁN THEO ACQ BANK" in title_cell

    # ======== CHECK 4: Check header mapping ========
    header_row = [cell.value for cell in ws[2]]
    assert "STT" in header_row        # NO -> STT
    assert "Phiên HT" in header_row   # RPT_SESSION -> Phiên HT
    assert "ACQ BANK" in header_row

    # ======== CHECK 5: Export thành công ========
    assert exporter.is_successful is True